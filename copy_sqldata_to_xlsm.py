import pandas as pd
import openpyxl
import sys
import time
df = pd.read_csv(sys.argv[2])#,dtype=str)
wb = openpyxl.Workbook()
s = wb.worksheets[0]

for i in range(1,len(df.columns)+1):
    for j in range(1,len(df)+1):
        if i ==5:
            s.cell(column=i,row=j).value="".join(list([val for val in list(df.iloc[:,i-1])[j-1] if val.isalnum()]))#converting a column in a dataframe to list- list(df[column_index\columnname])
        else:
            s.cell(column=i,row=j).value=list(df.iloc[:, i - 1])[j - 1]
        
wb.save(sys.argv[1].replace("Milestone.xlsm","Milestone_xl.xlsx"))
wb.close()
time.sleep(15)
print("success")