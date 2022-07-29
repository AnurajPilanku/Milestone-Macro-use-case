#P Anuraj
#SMO Milestone
import openpyxl
from datetime import datetime
import string
from dateutil import parser
import sys

wb=openpyxl.load_workbook(sys.argv[1])#r"C:\Users\2040664\anuraj\SMO\sacin.xlsx")
s=wb.active

for c in range(1,s.max_column+1):
    for r in range(2,s.max_row+1):
        if type(s[string.ascii_uppercase[c-1]+str(r)].value) in [str]:
            s[string.ascii_uppercase[c - 1] + str(r)].value=s[string.ascii_uppercase[c-1]+str(r)].value.strip()
        if c in [7,8]:
            d=s[string.ascii_uppercase[c-1]+str(r)].value
            g=parser.parse(d)
            s[string.ascii_uppercase[c - 1] + str(r)].value=g
            print(type(g))
wb.save(sys.argv[1])#r"C:\Users\2040664\anuraj\SMO\sacin.xlsx")


