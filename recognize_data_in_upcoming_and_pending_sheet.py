import openpyxl
import sys

up=openpyxl.load_workbook(sys.argv[1])
us=up.active
pe=openpyxl.load_workbook(sys.argv[2])
ps=pe.active

if us.cell(row=2,column=1).value in [None,"None"," ",""]:
    if us.cell(row=2, column=2).value in [None, "None", " ", ""]:
        print("failure")
    else:
        print("success")
else:
    print("success")
